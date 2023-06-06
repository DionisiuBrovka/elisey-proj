import random
import openpyxl
import datetime
from pathlib import Path
from dateutil.relativedelta import relativedelta
from ..models import Smena,Vospitanik,Otryad

BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLS_DIR =  Path('data/templs/')
MEDIA_DIR = BASE_DIR.parent / Path('media/')

def export_obhchiu(query):

    CUR_DATETIME = datetime.datetime.now()
    FILE_PATH_REL = f"obhchie/otchet_{str(CUR_DATETIME.day)}.{str(CUR_DATETIME.month)}.{str(CUR_DATETIME.year)}_{str(random.randint(10000,99999))}.xlsx"

    doc = openpyxl.load_workbook(BASE_DIR / TEMPLS_DIR / Path('obshiy.xlsx'))

    ws = doc.active

    start_row = 11
    start_col = 2

    row_n = start_row
    for vospitanik in Vospitanik.objects.filter(smena__id = query.id).order_by('-second_name'):
        ws.cell(row = start_row, column=2).value = f"{vospitanik.order_state} №{vospitanik.order_number}"
        ws.cell(row = start_row, column=3).value = f"{vospitanik.second_name} {vospitanik.first_name} {vospitanik.third_name}"

        bd = vospitanik.birthdate.strftime("%d.%m.%Y")
        ws.cell(row = start_row, column=4).value = f"{bd}, {vospitanik.ped_zav_full}, {vospitanik.state} класс"
        
        full_year = datetime.date.today() - vospitanik.birthdate
        ws.cell(row = start_row, column=5).value = f"{str(full_year.days // 365)}"

        ws.cell(row = start_row, column=6).value = f"{vospitanik.live_place_full}, {vospitanik.phone}"

        ws.cell(row = start_row, column=7).value = f"{vospitanik.mother_name_full} {vospitanik.mother_work} {vospitanik.mother_post} {vospitanik.mother_phone}"
        ws.cell(row = start_row, column=8).value = f"{vospitanik.father_name_full} {vospitanik.father_work} {vospitanik.father_post} {vospitanik.father_phone}"
        ws.cell(row = start_row, column=9).value = f"{vospitanik.coment}"

        start_row += 1

    doc.save(MEDIA_DIR / Path(FILE_PATH_REL))
    return FILE_PATH_REL


def export_soc(query):

    CUR_DATETIME = datetime.datetime.now()
    FILE_PATH_REL = f"soc/otchet_{str(CUR_DATETIME.day)}.{str(CUR_DATETIME.month)}.{str(CUR_DATETIME.year)}_{str(random.randint(10000,99999))}.xlsx"

    doc = openpyxl.load_workbook(BASE_DIR / TEMPLS_DIR / Path('soc_haracterictika.xlsx'))

    ws = doc.active

    start_row = 13
    start_col = 3

    row_n = start_row
    for otryad in Otryad.objects.all():
        vospitaniki = Vospitanik.objects.filter(otryad__id = otryad.pk).filter(smena__id = query.id)
        print(vospitaniki)

        for v in vospitaniki:
            print(v.soc)

        ws.cell(row = start_row, column=3).value =  vospitaniki.count()
        ws.cell(row = start_row, column=4).value =  vospitaniki.filter(sex='m').count()
        ws.cell(row = start_row, column=5).value =  vospitaniki.filter(sex='w').count()
        ws.cell(row = start_row, column=6).value =  vospitaniki.filter(soc='do').count()
        ws.cell(row = start_row, column=7).value =  vospitaniki.filter(soc='vdd').count() 

        ws.cell(row = start_row, column=8).value =  vospitaniki.filter(soc='di').count()
        ws.cell(row = start_row, column=9).value =  vospitaniki.filter(soc='ipr').count()
        ws.cell(row = start_row, column=10).value =  vospitaniki.filter(soc='sop').count() 
        ws.cell(row = start_row, column=11).value =  vospitaniki.filter(soc='opfr').count()

        ws.cell(row = start_row, column=12).value =  vospitaniki.filter(soc_fam='m').count()

        ws.cell(row = start_row, column=13).value =  vospitaniki.filter(soc_fam='np').count()
        ws.cell(row = start_row, column=14).value =  vospitaniki.filter(soc_fam='ri').count() 
        ws.cell(row = start_row, column=15).value =  vospitaniki.filter(soc_fam='rchz').count() 
        ws.cell(row = start_row, column=16).value =  vospitaniki.filter(soc_fam='b').count() 
        ws.cell(row = start_row, column=17).value =  vospitaniki.filter(soc_fam='os').count() 

        ws.cell(row = start_row, column=18).value =  vospitaniki.filter(soc_fam='ps').count()
        ws.cell(row = start_row, column=19).value =  vospitaniki.filter(soc_fam='rvdd').count()


        start_row += 1

    doc.save(MEDIA_DIR / Path(FILE_PATH_REL))
    return FILE_PATH_REL


def export_uchet_putevok(query):

    CUR_DATETIME = datetime.datetime.now()
    FILE_PATH_REL = f"put/otchet_{str(CUR_DATETIME.day)}.{str(CUR_DATETIME.month)}.{str(CUR_DATETIME.year)}_{str(random.randint(10000,99999))}.xlsx"

    doc = openpyxl.load_workbook(BASE_DIR / TEMPLS_DIR / Path('uchet_putevok.xlsx'))

    ws = doc.active

    start_row = 11
    start_col = 1

    row_n = start_row
    for vospitanik in Vospitanik.objects.filter(smena__id = query.id).order_by('-second_name'):
        ws.cell(row = start_row, column=1).value = f"{vospitanik.order_number}"
        ws.cell(row = start_row, column=2).value = f"{vospitanik.order_state}"
        ws.cell(row = start_row, column=3).value = f"{vospitanik.second_name} {vospitanik.first_name} {vospitanik.third_name}"

        bd = vospitanik.birthdate.strftime("%d.%m.%Y")
        ws.cell(row = start_row, column=4).value = f"{bd}"
        
        ws.cell(row = start_row, column=5).value = f"{vospitanik.ped_zav_full}"
        ws.cell(row = start_row, column=6).value = f"{Smena.objects.get(id=query.id)}"

        ws.cell(row = start_row, column=7).value = f"Мать : {vospitanik.mother_name_full} {vospitanik.mother_work} {vospitanik.mother_post} Отец : {vospitanik.father_name_full} {vospitanik.father_work} {vospitanik.father_post}"
        ws.cell(row = start_row, column=8).value = f"{vospitanik.live_place_full}"
        ws.cell(row = start_row, column=9).value = f"Мать : {vospitanik.mother_phone}, Отец : {vospitanik.father_phone}"

        start_row += 1

    doc.save(MEDIA_DIR / Path(FILE_PATH_REL))
    return FILE_PATH_REL